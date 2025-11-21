#!/usr/bin/env python3

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def create_excel_template():
    """Create Excel template for asset migration"""
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create Assets sheet
    assets_ws = wb.create_sheet("Assets")
    
    # Asset headers
    asset_headers = [
        # Basic Info
        "Asset Name", "Asset Category", "Asset Sub Category", "Department", "Asset Type", "Tenant ID",
        "Purchase Date", "Purchase Cost", "Current Value", "Description",
        
        # Address Info  
        "Address Line 1", "Address Line 2", "City", "Pincode", "Ward", "Zone", "Locality", "Division", "District",
        
        # Movable Asset Fields
        "Serial Number", "Model", "Manufacturer", "Warranty Expiry", "Condition", 
        "Maintenance Schedule", "Insurance Policy",
        
        # Immovable Asset Fields
        "Plot Area", "Built Up Area", "Floor Number", "Room Number", "Property Type", "Construction Year"
    ]
    
    # Add headers to Assets sheet
    for col, header in enumerate(asset_headers, 1):
        cell = assets_ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add sample data for Assets
    sample_assets = [
        [
            "Laptop Dell 5520", "IT Equipment", "Laptop", "IT Department", "MOVABLE", "pg.citya",
            "15/01/2023", 45000, 40000, "Dell Laptop for office use",
            "123 Main Street", "Near City Center", "Pune", 411001, "Ward 1", "Zone A", "Koregaon Park", "Pune Division", "Pune District",
            "DL123456", "Dell Inspiron 5520", "Dell Inc", "31/12/2025", "GOOD", "Quarterly", "POL123456",
            "", "", "", "", "", ""
        ],
        [
            "Office Building", "Infrastructure", "Building", "Administration", "IMMOVABLE", "pg.citya", 
            "01/04/2020", 5000000, 4500000, "Main office building",
            "456 Business District", "Sector 15", "Mumbai", 400001, "Ward 5", "Zone B", "Andheri East", "Mumbai Division", "Mumbai District",
            "", "", "", "", "", "", "",
            "2000 sqft", "1800 sqft", "Ground Floor", "101", "Office Building", "2020"
        ],
        [
            "Desktop Computer", "IT Equipment", "Desktop", "Accounts", "MOVABLE", "pg.citya",
            "10/03/2023", 35000, 32000, "Desktop for accounting work",
            "789 Finance Block", "", "Delhi", 110001, "Ward 3", "Zone C", "CP", "Delhi Division", "Delhi District",
            "DT789012", "HP EliteDesk", "HP Inc", "10/03/2026", "EXCELLENT", "Half-yearly", "POL789012",
            "", "", "", "", "", ""
        ]
    ]
    
    for row_idx, row_data in enumerate(sample_assets, 2):
        for col_idx, value in enumerate(row_data, 1):
            assets_ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Inventory sheet
    inventory_ws = wb.create_sheet("Inventory")
    
    # Inventory headers
    inventory_headers = [
        "Asset Reference", "Quantity", "Unit", "Supplier", "Purchase Order", "Invoice Number"
    ]
    
    # Add headers to Inventory sheet
    for col, header in enumerate(inventory_headers, 1):
        cell = inventory_ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add sample inventory data (using Excel row references)
    sample_inventory = [
        ["TEMP_AST_0002", 1, "NOS", "Dell India", "PO2023001", "INV2023001"],
        ["TEMP_AST_0004", 1, "NOS", "HP India", "PO2023002", "INV2023002"]
    ]
    
    for row_idx, row_data in enumerate(sample_inventory, 2):
        for col_idx, value in enumerate(row_data, 1):
            inventory_ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust column widths
    for ws in [assets_ws, inventory_ws]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save template
    template_path = "templates/asset_migration_template.xlsx"
    wb.save(template_path)
    print(f"Excel template created: {template_path}")
    
    # Create instructions sheet
    instructions_ws = wb.create_sheet("Instructions", 0)
    
    instructions = [
        ["ASSET MIGRATION TEMPLATE - INSTRUCTIONS"],
        [""],
        ["MANDATORY FIELDS:"],
        ["- Asset Name, Asset Category, Department, Asset Type, Tenant ID"],
        ["- For MOVABLE assets: Serial Number, Model"],
        ["- For IMMOVABLE assets: Plot Area, Property Type"],
        [""],
        ["ASSET TYPE VALUES:"],
        ["- MOVABLE (for laptops, desktops, vehicles, etc.)"],
        ["- IMMOVABLE (for buildings, land, infrastructure)"],
        [""],
        ["DATE FORMAT:"],
        ["- DD/MM/YYYY or DD-MM-YYYY or YYYY-MM-DD"],
        [""],
        ["CONDITION VALUES:"],
        ["- EXCELLENT, GOOD, FAIR, POOR"],
        [""],
        ["UNIT VALUES (for Inventory):"],
        ["- NOS, KG, LITER, METER, SQFT"],
        [""],
        ["NOTES:"],
        ["- Fill only relevant fields based on Asset Type"],
        ["- Movable assets use Serial Number, Model, etc."],
        ["- Immovable assets use Plot Area, Floor Number, etc."],
        ["- Inventory sheet is optional"],
        ["- Asset Reference should be TEMP_AST_XXXX format"],
        ["- Migration will auto-map to actual Asset IDs"]
    ]
    
    for row_idx, instruction in enumerate(instructions, 1):
        cell = instructions_ws.cell(row=row_idx, column=1, value=instruction[0])
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif "MANDATORY" in instruction[0] or "ASSET TYPE" in instruction[0] or "DATE FORMAT" in instruction[0]:
            cell.font = Font(bold=True)
    
    instructions_ws.column_dimensions['A'].width = 50
    
    wb.save(template_path)
    print(f"Excel template with instructions created: {template_path}")

if __name__ == "__main__":
    create_excel_template()